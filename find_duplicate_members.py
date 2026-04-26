
import base64
import os
import sys
import time
from collections import defaultdict
from datetime import datetime, timezone

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def load_env(path=".env"):
    """Load key=value pairs from a .env file into os.environ (won't overwrite)."""
    try:
        with open(path) as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, val = line.partition("=")
                key = key.strip()
                val = val.strip().strip('"').strip("'")
                os.environ.setdefault(key, val)
    except FileNotFoundError:
        pass

load_env(os.path.join(os.path.dirname(__file__), "pnaa/.env"))

WA_API_KEY    = os.environ.get("WILD_APRICOT_API_KEY")
WA_ACCOUNT_ID = os.environ.get("WILD_APRICOT_ACCOUNT_ID")

if not WA_API_KEY or not WA_ACCOUNT_ID:
    sys.exit(
        "ERROR: Missing Wild Apricot API key.\n"
        "  Add to pnaa/.env:\n"
        "    WILD_APRICOT_API_KEY=your_api_key\n"
        "    WILD_APRICOT_ACCOUNT_ID=213319\n\n"
        "  Get the API key from:\n"
        "  WA Admin → Settings → Applications → Authorized Applications\n"
    )



def get_wa_token(api_key: str) -> str:
    credentials = base64.b64encode(f"APIKEY:{api_key}".encode()).decode()
    resp = requests.post(
        "https://oauth.wildapricot.org/auth/token",
        headers={
            "Authorization": f"Basic {credentials}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        data="grant_type=client_credentials&scope=auto",
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def fetch_all_contacts(token: str, account_id: str) -> list[dict]:
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    base_url = f"https://api.wildapricot.org/v2/accounts/{account_id}/contacts"

    print("Requesting contacts from Wild Apricot…")
    resp = requests.get(base_url, headers=headers, timeout=60)
    resp.raise_for_status()
    data = resp.json()

    result_url = data.get("ResultUrl")
    if result_url and data.get("State") != "Complete":
        print("Waiting for WA async job to complete…")
        for attempt in range(96):          # up to 8 minutes
            time.sleep(5)
            poll = requests.get(result_url, headers=headers, timeout=30)
            poll.raise_for_status()
            data = poll.json()
            if data.get("State") == "Complete":
                print(f"  Done after ~{(attempt + 1) * 5}s")
                break
        else:
            sys.exit("ERROR: WA async job timed out.")

    total = data.get("ResultCount", 0)
    print(f"Total contacts: {total}")

    PAGE_SIZE = 100
    base = result_url or base_url
    contacts: list[dict] = []
    skip = 0

    while True:
        sep = "&" if "?" in base else "?"
        page_url = f"{base}{sep}$top={PAGE_SIZE}&$skip={skip}"
        page_resp = requests.get(page_url, headers=headers, timeout=60)
        if not page_resp.ok:
            print(f"  Warning: page request failed at skip={skip}: {page_resp.status_code}")
            break
        page_data = page_resp.json()
        batch = page_data.get("Contacts", [])
        if not batch:
            break
        contacts.extend(batch)
        skip += len(batch)
        print(f"  Fetched {skip}/{total}", end="\r")
        if len(batch) < PAGE_SIZE:
            break

    print(f"\nFetched {len(contacts)} contacts.")
    return contacts



def field_value(field_values: list[dict], name: str) -> str:
    for f in field_values:
        if f.get("FieldName") == name:
            v = f.get("Value")
            if v is None:
                return ""
            if isinstance(v, dict) and "Label" in v:
                return v["Label"]
            return str(v)
    return ""

def chapter_name(field_values: list[dict]) -> str:
    for f in field_values:
        if "Chapter" in (f.get("FieldName") or ""):
            v = f.get("Value")
            if v is None:
                continue
            label = v.get("Label", "") if isinstance(v, dict) else str(v)
            if label:
                return label
    return ""

def map_contact(contact: dict) -> dict | None:
    fv = contact.get("FieldValues", [])

    if field_value(fv, "Archived").lower() in ("true", "1", "yes"):
        return None

    renewal = field_value(fv, "Renewal due")
    try:
        renewal_dt = datetime.fromisoformat(renewal.replace("Z", "")) if renewal else None
        is_active = renewal_dt is not None and renewal_dt >= datetime.now()
    except (ValueError, TypeError):
        is_active = False

    level_obj = contact.get("MembershipLevel")
    level = level_obj.get("Name", "") if isinstance(level_obj, dict) else ""

    first = str(contact.get("FirstName") or "").strip()
    last  = str(contact.get("LastName")  or "").strip()
    name  = f"{first} {last}".strip()

    return {
        "Name":             name,
        "First Name":       first,
        "Last Name":        last,
        "Email":            str(contact.get("Email") or ""),
        "Member ID":        field_value(fv, "Member ID") or str(contact.get("Id", "")),
        "Membership Level": level,
        "Active Status":    "Active" if is_active else "Lapsed",
        "Renewal Due":      renewal,
        "Chapter":          chapter_name(fv),
        "Region":           field_value(fv, "PNAA Region"),
        "Education":        field_value(fv, "Highest Level of Education"),
        "WA Contact ID":    str(contact.get("Id", "")),
    }

# ── Find duplicates 

def find_duplicates(members: list[dict]) -> list[dict]:
    by_name: dict[str, list[dict]] = defaultdict(list)
    for m in members:
        if m["Name"]:
            by_name[m["Name"]].append(m)

    dupes: list[dict] = []
    for name, group in sorted(by_name.items()):
        if len(group) > 1:
            for m in group:
                dupes.append(m)

    print(f"Found {len(set(m['Name'] for m in dupes))} duplicate names ({len(dupes)} total rows).")
    return dupes

#excel

COLUMNS = [
    "Name", "First Name", "Last Name", "Email",
    "Member ID", "WA Contact ID", "Membership Level",
    "Active Status", "Renewal Due", "Chapter", "Region", "Education",
]

HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")   
ACTIVE_FILL   = PatternFill("solid", fgColor="D6F0E0")  
LAPSED_FILL   = PatternFill("solid", fgColor="FDE8E8")  
ALT_FILL      = PatternFill("solid", fgColor="F0F4FF")   

def write_excel(dupes: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duplicate Members"

    
    ws.append(COLUMNS)
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    
    ws.freeze_panes = "A2"

    
    group_toggle = False
    current_name = None
    for member in dupes:
        if member["Name"] != current_name:
            current_name = member["Name"]
            group_toggle = not group_toggle

        row_data = [member.get(col, "") for col in COLUMNS]
        ws.append(row_data)
        row_idx = ws.max_row

        base_fill = ALT_FILL if group_toggle else None
        status_fill = ACTIVE_FILL if member.get("Active Status") == "Active" else LAPSED_FILL

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == "Active Status":
                cell.fill = status_fill
            elif base_fill:
                cell.fill = base_fill

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(output_path)
    print(f"Saved: {output_path}")


def main():
    token    = get_wa_token(WA_API_KEY)
    contacts = fetch_all_contacts(token, WA_ACCOUNT_ID)

    members = [m for c in contacts if (m := map_contact(c)) is not None]
    print(f"Mapped {len(members)} non-archived members.")

    dupes = find_duplicates(members)
    if not dupes:
        print("No duplicate names found — nothing to export.")
        return

    output = os.path.join(os.path.dirname(__file__), "duplicate_members.xlsx")
    write_excel(dupes, output)

if __name__ == "__main__":
    main()
